#!/usr/bin/env python3
"""
微信聊天记录解析器

支持的导出格式：
1. WechatExporter 导出的 txt 文件（格式：时间 发送人: 内容）
2. WechatExporter 导出的 html 文件
3. 其他微信备份工具导出的 txt/csv

用法：
    python wechat_parser.py --file chat.txt --target "小美" --output output.txt
    python wechat_parser.py --file chat.html --target "小美" --output output.txt
"""

import re
import sys
import csv
import json
import argparse
from pathlib import Path
from html.parser import HTMLParser


class WechatHTMLParser(HTMLParser):
    """解析 WechatExporter 导出的 HTML 格式"""

    def __init__(self, target_name: str, bidirectional: bool = False):
        super().__init__()
        self.target_name = target_name
        self.bidirectional = bidirectional
        self.messages = []
        self._current_sender = ""
        self._current_time = ""
        self._current_content = []
        self._in_sender = False
        self._in_content = False
        self._in_time = False

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        cls = attrs_dict.get("class", "")

        if "sender" in cls:
            self._in_sender = True
        elif "content" in cls or "message-text" in cls:
            self._in_content = True
        elif "time" in cls or "timestamp" in cls:
            self._in_time = True

    def handle_endtag(self, tag):
        if self._in_sender:
            self._in_sender = False
        elif self._in_content:
            self._in_content = False
            content = "".join(self._current_content).strip()
            should_keep = self.bidirectional or self.target_name in self._current_sender
            if content and should_keep:
                self.messages.append({
                    "sender": self._current_sender,
                    "content": content,
                    "timestamp": self._current_time,
                })
            self._current_content = []
        elif self._in_time:
            self._in_time = False

    def handle_data(self, data):
        if self._in_sender:
            self._current_sender = data.strip()
        elif self._in_content:
            self._current_content.append(data)
        elif self._in_time:
            self._current_time = data.strip()


def parse_wechat_txt(file_path: str, target_name: str, bidirectional: bool = False) -> list[dict]:
    """解析微信导出的 TXT 格式消息"""
    messages = []

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="gbk", errors="replace") as f:
            lines = f.readlines()

    # 匹配常见格式：
    # 2024-01-01 10:00:00 小美: 消息内容
    # 2024/01/01 10:00 小美：消息内容
    pattern = re.compile(
        r"^(?P<time>\d{4}[-/]\d{1,2}[-/]\d{1,2}[\s\d:]*)\s+(?P<sender>.+?)[:：]\s*(?P<content>.+)$"
    )

    current_msg = None

    def _should_keep(sender: str) -> bool:
        if bidirectional:
            return True
        return target_name in sender

    for line in lines:
        line = line.rstrip("\n")
        if not line.strip():
            continue

        m = pattern.match(line)
        if m:
            # 保存上一条多行消息
            if current_msg and _should_keep(current_msg["sender"]):
                messages.append(current_msg)

            current_msg = {
                "sender": m.group("sender").strip(),
                "content": m.group("content").strip(),
                "timestamp": m.group("time").strip(),
            }
        elif current_msg:
            # 多行消息，追加到当前消息
            current_msg["content"] += "\n" + line

    # 最后一条
    if current_msg and _should_keep(current_msg["sender"]):
        messages.append(current_msg)

    # 过滤系统消息和媒体占位符
    filtered = []
    skip_patterns = [
        "[图片]", "[文件]", "[撤回了一条消息]", "[语音]", "[视频]",
        "[表情]", "[位置]", "[名片]", "[链接]", "[红包]", "[转账]",
        "<img", "<video", "<audio",
    ]
    for msg in messages:
        content = msg["content"].strip()
        if not content:
            continue
        if any(p in content for p in skip_patterns):
            continue
        filtered.append(msg)

    return filtered


def parse_wechat_csv(file_path: str, target_name: str, bidirectional: bool = False) -> list[dict]:
    """解析 CSV 格式的微信聊天记录"""
    messages = []

    try:
        f = open(file_path, "r", encoding="utf-8-sig")
        f.read(1)
        f.seek(0)
    except UnicodeDecodeError:
        f = open(file_path, "r", encoding="gbk", errors="replace")
    with f:
        reader = csv.DictReader(f)
        for row in reader:
            sender = (
                row.get("sender") or row.get("发送人") or
                row.get("昵称") or row.get("from") or row.get("NickName") or ""
            )
            content = (
                row.get("content") or row.get("内容") or
                row.get("message") or row.get("Message") or ""
            )
            timestamp = (
                row.get("timestamp") or row.get("时间") or
                row.get("time") or row.get("StrTime") or ""
            )

            if not bidirectional and target_name and target_name not in str(sender):
                continue
            if not content.strip():
                continue

            messages.append({
                "sender": str(sender),
                "content": str(content).strip(),
                "timestamp": str(timestamp),
            })

    return messages


def preview_xlsx(file_path: str, max_rows: int = 15) -> str:
    """预览 xlsx 文件的前 N 行原始内容，供 Claude 判断表头位置"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "错误：解析 xlsx 需要 openpyxl，请运行 pip install openpyxl"

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    lines = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = [str(v) if v is not None else "" for v in row]
        lines.append(f"第{i + 1}行: {' | '.join(cells)}")

    wb.close()
    return "\n".join(lines)


def parse_wechat_xlsx(file_path: str, target_name: str, header_row: int,
                       bidirectional: bool = False) -> list[dict]:
    """解析 XLSX 格式的微信聊天记录（WeFlow 等工具导出）

    Args:
        file_path: xlsx 文件路径
        target_name: 目标人物姓名
        header_row: 表头所在行号（1-based），由 Claude 通过 preview 判断
        bidirectional: 是否保留双向对话
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("错误：解析 xlsx 需要 openpyxl，请运行 pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if header_row < 1 or header_row > len(rows):
        print(f"错误：header_row={header_row} 超出范围（共 {len(rows)} 行）", file=sys.stderr)
        return []

    # header_row 是 1-based，转为 0-based 索引
    header_idx = header_row - 1
    headers = [str(v).strip() if v else "" for v in rows[header_idx]]
    messages = []

    for row in rows[header_idx + 1:]:
        row_dict = dict(zip(headers, [str(v) if v else "" for v in row]))

        # 按优先级匹配列名
        sender = (
            row_dict.get("发送者身份") or row_dict.get("发送人") or
            row_dict.get("昵称") or row_dict.get("sender") or
            row_dict.get("from") or row_dict.get("NickName") or ""
        )
        content = (
            row_dict.get("内容") or row_dict.get("content") or
            row_dict.get("message") or row_dict.get("Message") or ""
        )
        timestamp = (
            row_dict.get("时间") or row_dict.get("timestamp") or
            row_dict.get("time") or row_dict.get("StrTime") or ""
        )

        # 跳过系统消息（撤回通知等）
        msg_type = row_dict.get("消息类型", "")
        if msg_type == "系统消息":
            continue

        if not bidirectional and target_name and target_name not in sender:
            continue
        if not content.strip() or content.strip() == "None":
            continue

        messages.append({
            "sender": sender,
            "content": content.strip(),
            "timestamp": timestamp,
        })

    return messages


def parse_wechat_html(file_path: str, target_name: str, bidirectional: bool = False) -> list[dict]:
    """解析 HTML 格式的微信聊天记录"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            html_content = f.read()
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="gbk", errors="replace") as f:
            html_content = f.read()

    parser = WechatHTMLParser(target_name, bidirectional=bidirectional)
    parser.feed(html_content)
    return parser.messages


def extract_key_content(messages: list[dict]) -> dict:
    """
    对消息进行分类提取：
    - 长消息（>50字）：可能包含心情、想法、重要表达
    - 情感类回复：包含情感关键词
    - 日常沟通：其他消息
    """
    long_messages = []
    emotional_messages = []
    daily_messages = []

    emotional_keywords = [
        "想你", "爱你", "喜欢", "讨厌", "生气", "难过", "开心", "高兴",
        "不开心", "委屈", "对不起", "分手", "在一起", "想见你", "好想",
        "心疼", "舍不得", "感动", "幸福", "孤独", "害怕", "担心",
        "吵架", "冷战", "和好", "原谅", "道歉", "伤心", "哭",
        "miss", "love", "sorry", "happy", "sad",
    ]

    for msg in messages:
        content = msg["content"]

        if len(content) > 50:
            long_messages.append(msg)
        elif any(kw in content for kw in emotional_keywords):
            emotional_messages.append(msg)
        else:
            daily_messages.append(msg)

    return {
        "long_messages": long_messages,
        "emotional_messages": emotional_messages,
        "daily_messages": daily_messages,
        "total_count": len(messages),
    }


def normalize_timestamp(ts: str) -> str:
    """归一化时间戳为 YYYY-MM-DD HH:MM:SS 格式"""
    if not ts:
        return ""
    # 已经是标准格式
    if re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", ts):
        return ts
    # YYYY-MM-DD HH:MM -> 补 :00
    if re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$", ts):
        return ts + ":00"
    # YYYY/MM/DD HH:MM:SS 或 YYYY/MM/DD HH:MM
    ts = ts.replace("/", "-")
    if re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$", ts):
        return ts + ":00"
    return ts


def normalize_sender(sender: str, target_name: str) -> str:
    """归一化发送人为 her/me"""
    if sender in ("me", "我"):
        return "me"
    if target_name in sender:
        return "her"
    return "me"


def check_sender_mapping(messages: list[dict], target_name: str):
    """检测 sender 映射是否清晰，不清晰时输出警告"""
    unique_senders = set(msg["sender"] for msg in messages)
    has_wo = "我" in unique_senders
    has_target = any(target_name in s for s in unique_senders)

    if len(unique_senders) > 2:
        print(f"⚠️ 检测到 {len(unique_senders)} 个不同的发送人：{unique_senders}", file=sys.stderr)
        print(f"   预期只有 2 个：'我' 和 '{target_name}'", file=sys.stderr)
        print(f"   请检查输出的 JSON 中 sender 字段是否正确（应只有 'her' 和 'me'）", file=sys.stderr)
    elif not has_wo and not has_target:
        print(f"⚠️ 未检测到 '我' 或 '{target_name}'，发送人：{unique_senders}", file=sys.stderr)
        print(f"   请确认聊天记录中的昵称是否正确", file=sys.stderr)
    elif not has_wo:
        print(f"⚠️ 未检测到 '我'，发送人：{unique_senders}", file=sys.stderr)
        print(f"   如果 '{unique_senders}' 中有你自己的昵称，归一化会将其标记为 'me'", file=sys.stderr)


def format_json_output(target_name: str, messages: list[dict]) -> str:
    """格式化为 JSON 输出，保留双向对话"""
    check_sender_mapping(messages, target_name)
    output = []
    for msg in messages:
        output.append({
            "sender": normalize_sender(msg["sender"], target_name),
            "content": msg["content"],
            "timestamp": normalize_timestamp(msg["timestamp"]),
        })
    return json.dumps(output, ensure_ascii=False, indent=2)


def format_output(target_name: str, extracted: dict) -> str:
    """格式化输出，供 AI 分析使用"""
    lines = [
        f"# 微信聊天记录提取结果",
        f"目标人物：{target_name}",
        f"总消息数：{extracted['total_count']}",
        "",
        "---",
        "",
        "## 长消息（心情/想法类，权重最高）",
        "",
    ]

    for msg in extracted["long_messages"]:
        ts = f"[{msg['timestamp']}] " if msg["timestamp"] else ""
        lines.append(f"{ts}{msg['content']}")
        lines.append("")

    lines += [
        "---",
        "",
        "## 情感类消息",
        "",
    ]

    for msg in extracted["emotional_messages"]:
        ts = f"[{msg['timestamp']}] " if msg["timestamp"] else ""
        lines.append(f"{ts}{msg['content']}")
        lines.append("")

    lines += [
        "---",
        "",
        "## 日常沟通（风格参考）",
        "",
    ]

    for msg in extracted["daily_messages"][:200]:
        ts = f"[{msg['timestamp']}] " if msg["timestamp"] else ""
        lines.append(f"{ts}{msg['content']}")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="解析微信聊天记录导出文件")
    parser.add_argument("--file", required=True, help="输入文件路径（.txt / .html / .csv / .xlsx）")
    parser.add_argument("--target", default="", help="目标人物姓名（只提取此人发出的消息）")
    parser.add_argument("--output", default=None, help="输出文件路径（默认打印到 stdout）")
    parser.add_argument("--format", choices=["md", "json"], default="md", help="输出格式：md（默认）或 json")
    parser.add_argument("--bidirectional", action="store_true", help="保留双向对话（仅 json 格式时默认开启）")
    parser.add_argument("--preview", action="store_true", help="预览 xlsx 前 15 行，供判断表头位置")
    parser.add_argument("--header-row", type=int, default=None, help="xlsx 表头所在行号（1-based），配合 --preview 使用")

    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"错误：文件不存在 {file_path}", file=sys.stderr)
        sys.exit(1)

    suffix = file_path.suffix.lower()

    # xlsx preview 模式
    if args.preview:
        if suffix != ".xlsx":
            print("错误：--preview 仅支持 .xlsx 文件", file=sys.stderr)
            sys.exit(1)
        result = preview_xlsx(str(file_path))
        if args.output:
            with open(args.output, "w", encoding="utf-8") as f:
                f.write(result)
        else:
            print(result)
        return

    # JSON 格式默认开启双向
    bidirectional = args.bidirectional or args.format == "json"

    if suffix in (".html", ".htm"):
        messages = parse_wechat_html(str(file_path), args.target, bidirectional=bidirectional)
    elif suffix == ".csv":
        messages = parse_wechat_csv(str(file_path), args.target, bidirectional=bidirectional)
    elif suffix == ".xlsx":
        if not args.header_row:
            print("错误：xlsx 文件需要指定 --header-row，请先用 --preview 查看表头位置", file=sys.stderr)
            sys.exit(1)
        messages = parse_wechat_xlsx(str(file_path), args.target,
                                      header_row=args.header_row,
                                      bidirectional=bidirectional)
    else:
        messages = parse_wechat_txt(str(file_path), args.target, bidirectional=bidirectional)

    if not messages:
        print(f"警告：未找到消息", file=sys.stderr)
        if not bidirectional:
            print("提示：请检查目标姓名是否与文件中的发送人名称一致", file=sys.stderr)

    if args.format == "json":
        output = format_json_output(args.target, messages)
    else:
        extracted = extract_key_content(messages)
        output = format_output(args.target, extracted)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(output)
        print(f"已输出到 {args.output}，共 {len(messages)} 条消息", file=sys.stderr)
    else:
        print(output)


if __name__ == "__main__":
    main()
